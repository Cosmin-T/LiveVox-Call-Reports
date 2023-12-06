#!/bin/bash
# process.py

import pandas as pd
import glob
import os
import openpyxl.utils
import subprocess
from openpyxl import load_workbook
from logic.util import *
from logic.logs import *

conf_log()

def reporting(file_output, file_name):
    data = []
    files = glob.glob(file_name)

    if not files:
        logging.info("No file found")
        return

    for f in files:
        try:
            logging.info(f'File: {f} found.')
            with open(f, 'r') as file:
                for line in file:
                    fields = line.strip().split(',')
                    data.append(fields)
            logging.info('Data appended')

        except Exception as e:
            logging.error(f'Error: {e}')
            return

    df = pd.DataFrame(data, columns=None,)
    logging.info('DataFrame created')

    phone_dialed = df.columns[2]
    df = df.drop_duplicates(subset=phone_dialed)
    logging.info('Duplicates removed')

    df.to_excel(file_output, index=False, header=None)
    logging.info('DataFrame Exported to Excel')


def apply_filters_to_excel(file_output):
    try:
        workbook = load_workbook(file_output)
        sheet = workbook.active

        if "Invalid" not in workbook.sheetnames:
            invalid_sheet = workbook.create_sheet("Invalid")
            logging.info('Creating Invalid Sheet')
        else:
            invalid_sheet = workbook["Invalid"]
            invalid_sheet.delete_rows(1, invalid_sheet.max_row)

        workbook.move_sheet("Invalid", -len(workbook.sheetnames))
        logging.info('Moving to 1st position')

        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        invalid_sheet.append([cell.value for cell in header_row])
        logging.info('Copying headers from Sheet1 to Invalid')

        lv_result_col_index = None
        for cell in header_row:
            if cell.value == 'LiveVox_Result':
                lv_result_col_index = cell.col_idx
                break
        logging.info('Found LiveVox_Result column')

        if lv_result_col_index is None:
            logging.error("The LiveVox_Result column was not found in the header row.")

        for row in sheet.iter_rows(min_row=2):
            lv_result_cell = row[lv_result_col_index - 1]
            if str(lv_result_cell.value).startswith('Invalid'):
                invalid_sheet.append([cell.value for cell in row])
        logging.info('Copied Invalid data to Invalid Sheet')

        max_column = sheet.max_column
        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(max_column)}1"

        workbook.save(file_output)
        logging.info('Saved Successfully')

    except Exception as e:
        logging.error(f'Error: {e}')
        return

def remove_file(file_name):
    try:
        files = glob.glob(file_name)
        for file in files:
            os.remove(file)
        logging.info(f'FIle {file_name} removed. d')

        applescript = f"""
        tell application "Finder"
            activate
            close every Finder window
            make new Finder window to POSIX file "{FOLDER_PATH}"
            tell application "System Events" to keystroke "t" using command down
            delay 0.5
            set target of Finder window 1 to POSIX file "{OUTLOOK_TEMPLATE}"
        end tell
        """

        subprocess.run(['osascript', '-e', applescript], check=True)
        logging.info('Folders Open Successfully\n')

    except Exception as e:
        logging.error(f'Error: {e}\n')